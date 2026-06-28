"""月休计划业务逻辑"""

from __future__ import annotations

import io
import json
from calendar import monthrange
from datetime import date, datetime

from flask_jwt_extended import get_jwt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from steeltech_db.extensions import db
from steeltech_db.models import MonthlyRest, MonthlyRestLock, Personnel, Role, RolePersonnel
from app.services.auth_scope import get_user_scope as resolve_user_scope, is_jwt_admin


def get_user_scope(personnel_id: str) -> dict:
    """
    获取当前用户可编辑的人员范围。
    返回: { personnelId, role, editablePersonnelIds, team }
      role: 'admin' | 'leader' | 'member'
    """
    return resolve_user_scope(personnel_id)


def list_monthly_rest(*, year: int | None = None, month: int | None = None) -> list[dict]:
    """查询月休记录，可按年月筛选"""
    query = MonthlyRest.query
    if year is not None:
        query = query.filter_by(year=year)
    if month is not None:
        query = query.filter_by(month=month)
    rows = query.order_by(MonthlyRest.personnel_id).all()
    return [row.to_dict() for row in rows]


def get_month_lock(*, year: int, month: int) -> dict:
    """获取某年月是否已锁定（定稿）"""
    row = MonthlyRestLock.query.filter_by(year=year, month=month).first()
    if row is None:
        return {"year": year, "month": month, "locked": False, "lockedBy": None, "lockedAt": None}
    payload = row.to_dict()
    payload["year"] = year
    payload["month"] = month
    return payload


def is_month_locked(*, year: int, month: int) -> bool:
    return MonthlyRestLock.query.filter_by(year=year, month=month).first() is not None


def _is_admin_personnel(personnel_id: str) -> bool:
    """检查某 personnel_id 是否为管理员（dev 账号在 JWT claim 中处理）"""
    admin_role = (
        db.session.query(RolePersonnel)
        .join(Role, Role.id == RolePersonnel.role_id)
        .filter(RolePersonnel.personnel_id == personnel_id, Role.code == "admin")
        .first()
    )
    return admin_role is not None


def save_monthly_rest(data: dict) -> tuple[dict | None, str | None, int]:
    """创建或更新一条月休记录"""
    record_id = data.get("id", "").strip()
    personnel_id = data.get("personnelId", "").strip()
    year = data.get("year")
    month = data.get("month")
    rest_days = data.get("restDays", [])

    if not record_id:
        return None, "记录ID不能为空", 400
    if not personnel_id:
        return None, "人员ID不能为空", 400
    if year is None or month is None:
        return None, "年份和月份不能为空", 400
    if not isinstance(rest_days, list):
        return None, "restDays 必须是数组", 400

    # 若当月已定稿，则禁止员工保存（管理员/dev 账号例外）
    try:
        claims = get_jwt()
    except RuntimeError:
        claims = {}
    editor_personnel_id = claims.get("personnel_id")
    if isinstance(year, int) and isinstance(month, int) and is_month_locked(year=year, month=month):
        if not is_jwt_admin():
            return None, f"{year}年{month}月已定稿，员工无法再修改", 403

    # 校验是否为周末
    for day_str in rest_days:
        try:
            d = datetime.strptime(day_str, "%Y-%m-%d").date()
        except ValueError:
            return None, f"日期格式无效: {day_str}", 400
        if d.year != year or d.month != month:
            return None, f"日期 {day_str} 不在 {year}年{month}月", 400
        if d.weekday() not in (5, 6):
            return None, f"{day_str} 不是周六或周日，不能设为月休", 400

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    existing = MonthlyRest.query.get(record_id)

    if existing:
        existing.personnel_id = personnel_id
        existing.year = year
        existing.month = month
        existing.rest_days = json.dumps(sorted(rest_days), ensure_ascii=False)
        existing.updated_at = now
    else:
        existing = MonthlyRest(
            id=record_id,
            personnel_id=personnel_id,
            year=year,
            month=month,
            rest_days=json.dumps(sorted(rest_days), ensure_ascii=False),
            created_at=now,
            updated_at=now,
        )
        db.session.add(existing)

    db.session.commit()
    return existing.to_dict(), None, 200


def batch_save_monthly_rest(
    records: list[dict], *, editor_personnel_id: str | None = None
) -> tuple[list[dict] | None, str | None, int]:
    """批量保存月休记录，仅保存当前用户有权修改的记录"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    saved: list[dict] = []

    # 若当月已定稿，则禁止员工保存（管理员仍可）
    if records:
        year = records[0].get("year")
        month = records[0].get("month")
        if isinstance(year, int) and isinstance(month, int) and is_month_locked(year=year, month=month):
            if editor_personnel_id:
                scope = get_user_scope(editor_personnel_id)
                if scope.get("role") != "admin":
                    return None, f"{year}年{month}月已定稿，员工无法再修改", 403
            else:
                return None, "无法识别当前用户", 401

    # 确定可编辑范围
    editable_ids: set[str] | None = None
    if editor_personnel_id:
        scope = get_user_scope(editor_personnel_id)
        editable_ids = set(scope["editablePersonnelIds"])

    for data in records:
        record_id = data.get("id", "").strip()
        personnel_id = data.get("personnelId", "").strip()
        year = data.get("year")
        month = data.get("month")
        rest_days = data.get("restDays", [])

        if not record_id or not personnel_id or year is None or month is None:
            continue
        if not isinstance(rest_days, list):
            continue

        # 权限检查：只能保存可编辑范围内的人员
        if editable_ids is not None and personnel_id not in editable_ids:
            continue

        # 仅保留周末
        valid_days: list[str] = []
        for day_str in rest_days:
            try:
                d = datetime.strptime(day_str, "%Y-%m-%d").date()
            except ValueError:
                continue
            if d.weekday() in (5, 6) and d.year == year and d.month == month:
                valid_days.append(day_str)

        existing = MonthlyRest.query.get(record_id)
        if existing:
            existing.rest_days = json.dumps(sorted(valid_days), ensure_ascii=False)
            existing.updated_at = now
        else:
            existing = MonthlyRest(
                id=record_id,
                personnel_id=personnel_id,
                year=year,
                month=month,
                rest_days=json.dumps(sorted(valid_days), ensure_ascii=False),
                created_at=now,
                updated_at=now,
            )
            db.session.add(existing)

        saved.append(existing.to_dict())

    db.session.commit()
    return saved, None, 200


def finalize_monthly_rest(
    records: list[dict], *, editor_personnel_id: str
) -> tuple[dict | None, str | None, int]:
    """
    管理员定稿：先保存批量记录，再锁定该年月，禁止员工修改。
    返回: { saved: [...], lock: {...} }
    """
    scope = get_user_scope(editor_personnel_id)
    if scope.get("role") != "admin":
        return None, "无权限，仅管理员可定稿", 403

    if not records:
        return None, "没有可保存的数据", 400

    year = records[0].get("year")
    month = records[0].get("month")
    if not isinstance(year, int) or not isinstance(month, int):
        return None, "year 和 month 必须为整数", 400

    saved, error, status = batch_save_monthly_rest(records, editor_personnel_id=editor_personnel_id)
    if error:
        return None, error, status

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = MonthlyRestLock.query.filter_by(year=year, month=month).first()
    if row is None:
        row = MonthlyRestLock(year=year, month=month, locked_by=editor_personnel_id, locked_at=now)
        db.session.add(row)
    else:
        row.locked_by = editor_personnel_id
        row.locked_at = now
    db.session.commit()

    return {"saved": saved, "lock": row.to_dict()}, None, 200


def delete_monthly_rest(record_id: str) -> tuple[dict | None, str | None, int]:
    """删除一条月休记录"""
    row = MonthlyRest.query.get(record_id)
    if row is None:
        return None, "记录不存在", 404
    db.session.delete(row)
    db.session.commit()
    return {"id": record_id}, None, 200


# ── Excel 导出常量 ──

_EXCEL_FONT = "宋体"
_EXCEL_SYMBOL = "☀"
_EXCEL_COMPANY = "ITSS"
_EXCEL_DEPT = "设备技术部"
_EXCEL_LEAVE_TYPE = "单休"

# 列映射：I=9 对应日期1, AM=39 对应日期31
_DATE_START_COL = 9  # column I
_DATE_END_COL = 39   # column AM (31 days max)


def _make_font(size: int = 9, bold: bool = False, color_theme: int | None = 1) -> Font:
    """创建宋体字体"""
    kwargs: dict = dict(name=_EXCEL_FONT, size=size, bold=bold)
    if color_theme is not None:
        from openpyxl.styles.colors import Color
        kwargs["color"] = Color(theme=color_theme)
    return Font(**kwargs)


_THIN_SIDE = Side(style="thin")


def _make_thin_border() -> Border:
    """创建四边细线边框"""
    return Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


def _make_partial_border(*, left: bool = False, right: bool = False, top: bool = False, bottom: bool = False) -> Border:
    """创建部分边框"""
    return Border(
        left=_THIN_SIDE if left else None,
        right=_THIN_SIDE if right else None,
        top=_THIN_SIDE if top else None,
        bottom=_THIN_SIDE if bottom else None,
    )


def _make_center_alignment(wrap_text: bool = False) -> Alignment:
    """创建居中、垂直居中对齐"""
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap_text)


def _apply_cell_style(cell, font: Font | None = None, alignment: Alignment | None = None, border: Border | None = None):
    """统一设置单元格样式"""
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def export_monthly_rest_excel(year: int, month: int) -> bytes:
    """
    生成月休计划 Excel 文件，返回字节流。
    格式参照 2026年4月调休汇总表（钢结构技术科）（定稿）.xlsx。
    """
    # ── 查询数据 ──
    records = (
        db.session.query(MonthlyRest, Personnel)
        .join(Personnel, MonthlyRest.personnel_id == Personnel.id)
        .filter(MonthlyRest.year == year, MonthlyRest.month == month, Personnel.status == "active")
        .order_by(Personnel.team, Personnel.name)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "调休汇总表"

    # ── 通用样式 ──
    header_font = _make_font(size=9, bold=True)
    title_font = _make_font(size=20, bold=True)
    normal_font = _make_font(size=9, bold=False)
    date_formula_font = _make_font(size=9, bold=False)
    year_month_font = _make_font(size=9, bold=True, color_theme=8)  # 深色文字
    thin_border = _make_thin_border()
    center_align = _make_center_alignment()
    center_wrap_align = _make_center_alignment(wrap_text=True)
    center_align_no_border = Alignment(horizontal="center", vertical="center")

    # ── 页面设置 ──
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.top = 0.55
    ws.page_margins.bottom = 1.0
    ws.page_margins.left = 0.39
    ws.page_margins.right = 0.24

    # ── 列宽 ──
    col_widths = {
        "A": 5.5, "B": 7.13, "C": 11.5, "D": 9.32, "E": 10.0,
        "F": 7.13, "G": 7.13, "H": 6.38,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    # 日期列 I~AM
    for col_idx in range(_DATE_START_COL, _DATE_END_COL + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 3.63
    ws.column_dimensions["AN"].width = 8.05
    ws.column_dimensions["AO"].width = 18.53
    ws.column_dimensions["AP"].width = 8.86

    # ── 标题行 (Row 1) ──
    title = f"{year}年{month}月调休汇总表"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=41)  # A1:AO1 (cols 1-41)
    title_cell = ws.cell(row=1, column=1, value=title)
    _apply_cell_style(title_cell, font=title_font, alignment=center_align)
    ws.row_dimensions[1].height = 30

    # ── 表头行 (Row 2) ──
    headers = [
        (1, "序号"), (2, "公司"), (3, "部门"), (4, "工号"), (5, "姓名"),
        (6, "岗位"), (7, "调休类别"), (8, "调休时段"),
    ]
    for col, text in headers:
        cell = ws.cell(row=2, column=col, value=text)
        _apply_cell_style(cell, font=header_font, alignment=center_align, border=thin_border)

    # I2:J2 合并 "符号"
    ws.merge_cells(start_row=2, start_column=9, end_row=2, end_column=10)
    cell = ws.cell(row=2, column=9, value="符号")
    _apply_cell_style(cell, font=header_font, alignment=center_align, border=thin_border)
    # K2: ☀ 符号
    cell = ws.cell(row=2, column=11, value=_EXCEL_SYMBOL)
    _apply_cell_style(cell, font=_make_font(size=10, bold=False), alignment=center_align, border=thin_border)
    # L2:M2 合并 "调休"
    ws.merge_cells(start_row=2, start_column=12, end_row=2, end_column=13)
    cell = ws.cell(row=2, column=12, value="调休")
    _apply_cell_style(cell, font=header_font, alignment=center_align, border=thin_border)
    # N2:AI2 合并 (空, 作为日期区域的表头)
    ws.merge_cells(start_row=2, start_column=14, end_row=2, end_column=35)
    # AJ2:AM2: 年/月标题
    ymt_border = _make_partial_border(top=True, bottom=True)
    cell = ws.cell(row=2, column=36, value=year)
    _apply_cell_style(cell, font=year_month_font, alignment=center_align, border=ymt_border)
    cell = ws.cell(row=2, column=37, value="年")
    _apply_cell_style(cell, font=year_month_font, alignment=center_align, border=ymt_border)
    cell = ws.cell(row=2, column=38, value=month)
    _apply_cell_style(cell, font=year_month_font, alignment=center_align, border=ymt_border)
    cell = ws.cell(row=2, column=39, value="月")
    _apply_cell_style(cell, font=year_month_font, alignment=center_align, border=ymt_border)
    # AN2: 合计
    cell = ws.cell(row=2, column=40, value="合计")
    _apply_cell_style(cell, font=_make_font(size=9, bold=True, color_theme=None), alignment=center_align, border=_make_partial_border(left=True, top=True))
    # AO2: 备注
    cell = ws.cell(row=2, column=41, value="备注")
    _apply_cell_style(cell, font=_make_font(size=11, bold=True, color_theme=None), alignment=center_wrap_align, border=_make_partial_border(left=True, top=True))
    # AP2: 车间
    cell = ws.cell(row=2, column=42, value="车间")
    _apply_cell_style(cell, font=_make_font(size=12, bold=True, color_theme=None), alignment=center_align)

    ws.row_dimensions[2].height = 21

    # ── 日期表头行 (Row 3: 星期, Row 4: 日期数字) ──
    _, days_in_month = monthrange(year, month)
    weekday_names = ["一", "二", "三", "四", "五", "六", "日"]
    for day_idx in range(31):
        col = _DATE_START_COL + day_idx
        if day_idx < days_in_month:
            d = date(year, month, day_idx + 1)
            wday_cn = weekday_names[d.weekday()]  # 0=周一 → "一"
            # 周末用红色字体
            is_weekend = d.weekday() >= 5
            day_font = _make_font(size=9, bold=False, color_theme=1) if not is_weekend else Font(name=_EXCEL_FONT, size=9, bold=False, color="FF0000")
            wday_font = Font(name=_EXCEL_FONT, size=9, bold=False, color="FF0000") if is_weekend else _make_font(size=9, bold=False, color_theme=1)
            # Row 3: 星期几
            cell_w = ws.cell(row=3, column=col, value=wday_cn)
            _apply_cell_style(cell_w, font=wday_font, alignment=center_align, border=thin_border)
            # Row 4: 日期数字
            cell_d = ws.cell(row=4, column=col, value=day_idx + 1)
            _apply_cell_style(cell_d, font=day_font, alignment=center_align, border=thin_border)
        else:
            # 超出当月天数的列留空
            cell_w = ws.cell(row=3, column=col, value=None)
            _apply_cell_style(cell_w, font=date_formula_font, alignment=center_align, border=thin_border)
            cell_d = ws.cell(row=4, column=col, value=None)
            _apply_cell_style(cell_d, font=date_formula_font, alignment=center_align, border=thin_border)

    # AN3: 合计表头
    cell = ws.cell(row=3, column=40, value="合计")
    _apply_cell_style(cell, font=_make_font(size=9, bold=True, color_theme=None), alignment=center_align, border=_make_partial_border(left=True, top=True, right=True))
    ws.row_dimensions[3].height = 25
    ws.row_dimensions[4].height = 25

    # ── 合并表头区域 (Row 2-4) ──
    # A~H, AN, AO, AP 合并 rows 2-4
    merge_cols_2_to_4 = [1, 2, 3, 4, 5, 6, 7, 8, 40, 42]  # A-H, AN, AP
    for col in merge_cols_2_to_4:
        ws.merge_cells(start_row=2, start_column=col, end_row=4, end_column=col)
    # AO 合并 2-4
    ws.merge_cells(start_row=2, start_column=41, end_row=4, end_column=41)  # AO
    # AN 合并 3-4 (header row 2 stays separate)
    ws.merge_cells(start_row=3, start_column=40, end_row=4, end_column=40)

    # ── 数据行 ──
    data_start_row = 5
    for idx, (rest_record, person) in enumerate(records):
        seq = idx + 1
        row_am = data_start_row + idx * 2      # 上午行
        row_pm = data_start_row + idx * 2 + 1  # 下午行

        rest_days: list[str] = []
        try:
            days = json.loads(rest_record.rest_days) if isinstance(rest_record.rest_days, str) else rest_record.rest_days
            if isinstance(days, list):
                rest_days = days
        except (json.JSONDecodeError, TypeError):
            pass

        # 将 rest_days 转为 day-of-month 的集合
        rest_day_numbers: set[int] = set()
        for day_str in rest_days:
            try:
                d = datetime.strptime(day_str, "%Y-%m-%d").date()
                if d.year == year and d.month == month:
                    rest_day_numbers.add(d.day)
            except ValueError:
                pass

        # 合并单元格的列 (A~G, AN, AO, AP)
        merge_cols = [1, 2, 3, 4, 5, 6, 7, 40, 41, 42]
        for col in merge_cols:
            ws.merge_cells(start_row=row_am, start_column=col, end_row=row_pm, end_column=col)

        # A: 序号
        cell = ws.cell(row=row_am, column=1, value=seq)
        _apply_cell_style(cell, font=normal_font, alignment=center_align, border=thin_border)
        # B: 公司
        cell = ws.cell(row=row_am, column=2, value=_EXCEL_COMPANY)
        _apply_cell_style(cell, font=normal_font, alignment=center_align, border=thin_border)
        # C: 部门
        cell = ws.cell(row=row_am, column=3, value=person.workshop or _EXCEL_DEPT)
        _apply_cell_style(cell, font=normal_font, alignment=center_align, border=thin_border)
        # D: 工号
        cell = ws.cell(row=row_am, column=4, value=person.employee_no)
        _apply_cell_style(cell, font=normal_font, alignment=center_align, border=thin_border)
        # E: 姓名
        cell = ws.cell(row=row_am, column=5, value=person.name)
        _apply_cell_style(cell, font=normal_font, alignment=center_align, border=thin_border)
        # F: 岗位
        cell = ws.cell(row=row_am, column=6, value=person.position or "")
        _apply_cell_style(cell, font=normal_font, alignment=center_align, border=thin_border)
        # G: 调休类别
        cell = ws.cell(row=row_am, column=7, value=_EXCEL_LEAVE_TYPE)
        _apply_cell_style(cell, font=normal_font, alignment=center_align, border=thin_border)
        # H: 上午
        cell_am = ws.cell(row=row_am, column=8, value="上午")
        _apply_cell_style(cell_am, font=normal_font, alignment=center_align, border=thin_border)
        # H: 下午
        cell_pm = ws.cell(row=row_pm, column=8, value="下午")
        _apply_cell_style(cell_pm, font=normal_font, alignment=center_align, border=thin_border)

        # I~AM: 日期列，休息日标注 ☀
        for day_num in range(1, 32):
            col = _DATE_START_COL + day_num - 1
            is_rest = day_num in rest_day_numbers
            # 上午行
            cell_a = ws.cell(row=row_am, column=col, value=_EXCEL_SYMBOL if is_rest else None)
            _apply_cell_style(cell_a, font=normal_font, alignment=center_align, border=thin_border)
            # 下午行
            cell_p = ws.cell(row=row_pm, column=col, value=_EXCEL_SYMBOL if is_rest else None)
            _apply_cell_style(cell_p, font=normal_font, alignment=center_align, border=thin_border)

        # AN: 合计公式 =COUNTIF(I{am}:AM{pm},"☀")/2
        i_col = get_column_letter(_DATE_START_COL)
        am_col = get_column_letter(_DATE_END_COL)
        count_formula = f'=COUNTIF({i_col}{row_am}:{am_col}{row_pm},"{_EXCEL_SYMBOL}")/2'
        cell = ws.cell(row=row_am, column=40, value=count_formula)
        _apply_cell_style(cell, font=normal_font, alignment=center_align, border=thin_border)

        # AO: 备注 (空)
        cell = ws.cell(row=row_am, column=41, value=None)
        _apply_cell_style(cell, font=normal_font, alignment=center_align, border=_make_partial_border(left=True))

        # AP: 车间
        cell = ws.cell(row=row_am, column=42, value=person.team or "")
        _apply_cell_style(cell, font=normal_font, alignment=center_align)

        # 设置行高
        ws.row_dimensions[row_am].height = 25
        ws.row_dimensions[row_pm].height = 25

        # 为合并单元格的第二行也设置边框（避免合并后边框丢失）
        for col in range(1, 43):
            cell_pm_val = ws.cell(row=row_pm, column=col)
            if cell_pm_val.value is None and col not in merge_cols + [8]:
                _apply_cell_style(cell_pm_val, border=thin_border)

    # 如果无数据，至少保留空行框架
    if not records:
        for r in range(5, 7):
            ws.row_dimensions[r].height = 25
            for col in range(1, 43):
                cell = ws.cell(row=r, column=col)
                _apply_cell_style(cell, border=thin_border)

    # ── 冻结窗格 ──
    ws.freeze_panes = "A5"

    # ── 输出为字节流 ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
