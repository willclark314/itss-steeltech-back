"""钢构发图计划业务逻辑"""

from __future__ import annotations

import io
import json
import re
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from steeltech_db.extensions import db
from steeltech_db.models import DrawingIssuePlan, SystemSetting

_PLAN_TITLE_PATTERN = re.compile(r"(\d{4})年(\d{1,2})月")
_HEADER_MARKERS = ("序号", "项目名称", "项目编号")
_TOTAL_MARKERS = ("合计", "总计")
_DISPLAY_SETTINGS_KEY = "drawing_issue_plan_display"
_MIN_APPROACHING_DAYS = 0
_MAX_APPROACHING_DAYS = 30
_DEFAULT_APPROACHING_DAYS = 7


def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _build_record_id(year: int, month: int, seq: int) -> str:
    return f"{year}_{month:02d}_{seq:03d}"


def _normalize_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _normalize_weight(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_weight_header_in_tons(header: str) -> bool:
    normalized = (header or "").upper()
    if "KG" in normalized:
        return False
    return "吨" in header or "TON" in normalized


def _normalize_imported_weight(value: Any, *, in_tons: bool) -> float | None:
    weight = _normalize_weight(value)
    if weight is None:
        return None
    return weight * 1000 if in_tons else weight


def _kg_to_tons(value: float | None) -> float | None:
    if value is None:
        return None
    return value / 1000


def list_drawing_issue_plans(*, year: int | None = None, month: int | None = None) -> list[dict]:
    query = DrawingIssuePlan.query
    if year is not None:
        query = query.filter_by(year=year)
    if month is not None:
        query = query.filter_by(month=month)
    rows = query.order_by(
        DrawingIssuePlan.year.desc(),
        DrawingIssuePlan.month.desc(),
        DrawingIssuePlan.seq.asc(),
    ).all()
    return [row.to_dict() for row in rows]


def _validate_record(data: dict, *, require_id: bool = True) -> tuple[dict | None, str | None]:
    record_id = str(data.get("id", "")).strip()
    year = data.get("year")
    month = data.get("month")
    seq = data.get("seq")
    project_name = str(data.get("projectName", "")).strip()
    project_no = str(data.get("projectNo", "")).strip()
    component_weight_kg = _normalize_weight(data.get("componentWeightKg"))
    planned_issue_date = _normalize_date(data.get("plannedIssueDate"))
    actual_issue_date = _normalize_date(data.get("actualIssueDate"))
    material_prep_list_issue_date = _normalize_date(data.get("materialPrepListIssueDate"))
    remark = str(data.get("remark", "") or "").strip()

    if require_id and not record_id:
        return None, "记录ID不能为空"
    if year is None or month is None or seq is None:
        return None, "年份、月份和序号不能为空"
    try:
        year = int(year)
        month = int(month)
        seq = int(seq)
    except (TypeError, ValueError):
        return None, "年份、月份和序号必须为整数"
    if month < 1 or month > 12:
        return None, "月份必须在 1-12 之间"
    if seq < 1:
        return None, "序号必须大于 0"
    if not project_name:
        return None, "项目名称不能为空"
    if not project_no:
        return None, "项目编号不能为空"

    if not record_id:
        record_id = _build_record_id(year, month, seq)

    return {
        "id": record_id,
        "year": year,
        "month": month,
        "seq": seq,
        "project_name": project_name,
        "project_no": project_no,
        "component_weight_kg": component_weight_kg,
        "planned_issue_date": planned_issue_date,
        "actual_issue_date": actual_issue_date,
        "material_prep_list_issue_date": material_prep_list_issue_date,
        "remark": remark,
    }, None


def batch_replace_month_plans(records: list[dict]) -> tuple[list[dict] | None, str | None, int]:
    if not records:
        return None, "请求体不能为空", 400

    normalized: list[dict] = []
    year: int | None = None
    month: int | None = None

    for item in records:
        payload, error = _validate_record(item, require_id=False)
        if error or payload is None:
            return None, error, 400
        if year is None:
            year = payload["year"]
            month = payload["month"]
        elif payload["year"] != year or payload["month"] != month:
            return None, "批量保存的记录必须属于同一年月", 400
        normalized.append(payload)

    assert year is not None and month is not None
    now = _now_str()

    DrawingIssuePlan.query.filter_by(year=year, month=month).delete()
    saved: list[dict] = []
    for payload in sorted(normalized, key=lambda item: item["seq"]):
        record_id = _build_record_id(payload["year"], payload["month"], payload["seq"])
        row = DrawingIssuePlan(
            id=record_id,
            year=payload["year"],
            month=payload["month"],
            seq=payload["seq"],
            project_name=payload["project_name"],
            project_no=payload["project_no"],
            component_weight_kg=payload["component_weight_kg"],
            planned_issue_date=payload["planned_issue_date"],
            actual_issue_date=payload["actual_issue_date"],
            material_prep_list_issue_date=payload["material_prep_list_issue_date"],
            remark=payload["remark"],
            created_at=now,
            updated_at=now,
        )
        db.session.add(row)
        saved.append(row.to_dict())

    db.session.commit()
    return saved, None, 200


def _parse_title_year_month(title: str) -> tuple[int | None, int | None]:
    match = _PLAN_TITLE_PATTERN.search(title or "")
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def _is_header_row(values: list[Any]) -> bool:
    text = "".join(str(v or "") for v in values)
    return all(marker in text for marker in _HEADER_MARKERS)


def _is_total_row(values: list[Any]) -> bool:
    first = str(values[0] if values else "").strip()
    return any(marker in first for marker in _TOTAL_MARKERS)


def _find_header_index(headers: list[str], *markers: str) -> int | None:
    for index, header in enumerate(headers):
        if any(marker in header for marker in markers):
            return index
    return None


def _parse_import_row(values: list[Any], headers: list[str] | None, target_year: int, target_month: int, seq: int) -> dict | None:
    project_name_index = _find_header_index(headers or [], "项目名称") if headers else 1
    project_no_index = _find_header_index(headers or [], "项目编号") if headers else 2
    weight_index = _find_header_index(headers or [], "构件重量", "重量") if headers else 3
    prep_list_index = _find_header_index(headers or [], "备料清单") if headers else None
    planned_index = _find_header_index(headers or [], "计划发图") if headers else 4
    actual_index = _find_header_index(headers or [], "实际发图") if headers else None
    remark_index = _find_header_index(headers or [], "备注") if headers else 5

    if project_name_index is None:
        project_name_index = 1
    if project_no_index is None:
        project_no_index = 2
    if weight_index is None:
        weight_index = 3
    if planned_index is None:
        planned_index = 5 if prep_list_index is not None else 4
    if prep_list_index is None and headers is None:
        prep_list_index = 4
    if remark_index is None:
        if actual_index is not None:
            remark_index = 7 if prep_list_index is not None else 6
        else:
            remark_index = 6 if prep_list_index is not None else 5

    project_name = str(values[project_name_index] or "").strip() if len(values) > project_name_index else ""
    project_no = str(values[project_no_index] or "").strip() if len(values) > project_no_index else ""
    if not project_name and not project_no:
        return None

    seq_value = values[0] if values else seq
    parsed_seq = int(seq_value) if str(seq_value).strip().isdigit() else seq

    weight_header = headers[weight_index] if headers and weight_index is not None and len(headers) > weight_index else ""
    weight_in_tons = _is_weight_header_in_tons(weight_header)

    return {
        "year": target_year,
        "month": target_month,
        "seq": parsed_seq,
        "projectName": project_name,
        "projectNo": project_no,
        "componentWeightKg": _normalize_imported_weight(
            values[weight_index] if len(values) > weight_index else None,
            in_tons=weight_in_tons,
        ),
        "materialPrepListIssueDate": _normalize_date(
            values[prep_list_index] if prep_list_index is not None and len(values) > prep_list_index else None
        ),
        "plannedIssueDate": _normalize_date(values[planned_index] if len(values) > planned_index else None),
        "actualIssueDate": _normalize_date(values[actual_index] if actual_index is not None and len(values) > actual_index else None),
        "remark": str(values[remark_index] or "").strip() if len(values) > remark_index else "",
    }


def import_drawing_issue_plan_excel(
    file_bytes: bytes,
    *,
    year: int | None = None,
    month: int | None = None,
) -> tuple[list[dict] | None, str | None, int]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return None, f"无法读取 Excel 文件: {exc}", 400

    worksheet = workbook.active
    if worksheet is None:
        return None, "Excel 工作表为空", 400

    title_year, title_month = _parse_title_year_month(str(worksheet.cell(1, 1).value or ""))
    target_year = year or title_year
    target_month = month or title_month
    if target_year is None or target_month is None:
        return None, "无法识别计划年月，请在标题中包含“YYYY年M月”或手动指定 year/month", 400

    records: list[dict] = []
    headers: list[str] | None = None
    seq = 0
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        values = list(row) if row else []
        if not any(values):
            continue
        if _is_header_row(values):
            headers = [str(value or "").strip() for value in values]
            continue
        if _is_total_row(values):
            break

        seq += 1
        parsed = _parse_import_row(values, headers, target_year, target_month, seq)
        if parsed is None:
            continue
        records.append(parsed)

    if not records:
        return None, "未在 Excel 中找到有效细化发图计划数据", 400

    return batch_replace_month_plans(records)


def export_drawing_issue_plan_excel(year: int, month: int) -> bytes:
    records = (
        DrawingIssuePlan.query.filter_by(year=year, month=month)
        .order_by(DrawingIssuePlan.seq.asc())
        .all()
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "细化发图计划"

    title = f"设备技术部{year}年{month}月钢构细化发图计划"
    headers = ["序号", "项目名称", "项目编号", "构件重量/吨", "备料清单", "计划发图时间", "实际发图时间", "备注"]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    title_font = Font(name="宋体", size=14, bold=True)
    header_font = Font(name="宋体", size=11, bold=True)
    body_font = Font(name="宋体", size=11)

    worksheet.merge_cells("A1:H1")
    title_cell = worksheet["A1"]
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = center

    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    total_weight_kg = 0.0
    for index, record in enumerate(records, start=1):
        row_num = index + 2
        weight_kg = record.component_weight_kg or 0
        total_weight_kg += weight_kg
        values = [
            record.seq,
            record.project_name,
            record.project_no,
            _kg_to_tons(record.component_weight_kg),
            record.material_prep_list_issue_date,
            record.planned_issue_date,
            record.actual_issue_date,
            record.remark or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_num, column=col, value=value)
            cell.font = body_font
            cell.alignment = center if col in (1, 4, 5, 6, 7) else Alignment(vertical="center")
            cell.border = thin_border

    total_row = len(records) + 3
    worksheet.cell(row=total_row, column=1, value="合计").font = header_font
    total_cell = worksheet.cell(row=total_row, column=4, value=_kg_to_tons(total_weight_kg))
    total_cell.font = header_font
    total_cell.alignment = center

    worksheet.column_dimensions["A"].width = 8
    worksheet.column_dimensions["B"].width = 42
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 16
    worksheet.column_dimensions["E"].width = 16
    worksheet.column_dimensions["F"].width = 16
    worksheet.column_dimensions["G"].width = 16
    worksheet.column_dimensions["H"].width = 18

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _normalize_approaching_days(value: Any) -> int:
    try:
        days = int(value)
    except (TypeError, ValueError):
        days = _DEFAULT_APPROACHING_DAYS
    return max(_MIN_APPROACHING_DAYS, min(_MAX_APPROACHING_DAYS, days))


def get_display_settings() -> dict:
    row = SystemSetting.query.filter_by(key=_DISPLAY_SETTINGS_KEY).first()
    if row and row.value:
        try:
            parsed = json.loads(row.value)
            if isinstance(parsed, dict):
                return {
                    "approachingDays": _normalize_approaching_days(parsed.get("approachingDays")),
                }
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
    return {"approachingDays": _DEFAULT_APPROACHING_DAYS}


def save_display_settings(payload: dict) -> dict:
    if not isinstance(payload, dict):
        raise ValueError("请求体格式错误")

    normalized = {
        "approachingDays": _normalize_approaching_days(
            payload.get("approachingDays", _DEFAULT_APPROACHING_DAYS),
        ),
    }
    row = SystemSetting.query.filter_by(key=_DISPLAY_SETTINGS_KEY).first()
    now = _now_str()
    value = json.dumps(normalized, ensure_ascii=False)

    if row is None:
        row = SystemSetting(key=_DISPLAY_SETTINGS_KEY, value=value, updated_at=now)
        db.session.add(row)
    else:
        row.value = value
        row.updated_at = now

    db.session.commit()
    return normalized
